#!/usr/bin/env python3

import argparse
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def build_xlsx(year):
    """Build the ASR xlsx file with headers and formatting."""

    # Determine project root (parent of scripts directory)
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    # Define output path
    output_path = project_root / "input" / f"asr_{year}.xlsx"

    # Column definitions
    columns = [
        "title",
        "author(s)",
        "published__online_date",
        "article_url",
        "OnlineFirst (Y/N)",
        "volume",
        "issue",
        "in_scope(Y/NA)",
        "qualitative(Y/N)",
        "data(Y/N)",
        "code(Y/N)",
        "data + code",
        "neither",
        "data_gated(Y/N)",
        "data_source / apply_at",
        "package_location",
        "path_to_package",
        "coverage_checked",
        "notes"
    ]

    # Column width mapping
    column_widths = {
        "title": 62,
        "author(s)": 34,
        "article_url": 46,
        "notes": 40,
        "path_to_package": 34,
        "data_source / apply_at": 26,
        "package_location": 26,
    }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"asr_{year}"

    # Add headers to row 1
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)

        # Apply header formatting
        # Fill: #1F3864 solid
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

        # Font: bold, white, size 11 (default)
        cell.font = Font(bold=True, color="FFFFFF")

        # Alignment: centered, wrap text
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set row height for header row
    ws.row_dimensions[1].height = 34

    # Set column widths
    for col_num, column_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        width = column_widths.get(column_name, 16)  # Default to 16 for "others"
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at A2
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_path)
    print(f"Created {output_path}")

    return output_path

def verify_xlsx(year):
    """Verify the created xlsx file and print key details."""

    # Determine project root (parent of scripts directory)
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    # Load the workbook
    from openpyxl import load_workbook
    output_path = project_root / "input" / f"asr_{year}.xlsx"

    wb = load_workbook(output_path)
    ws = wb.active

    # Get sheet name
    print(f"\nSheet name: {ws.title}")

    # Get header list
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    # Get header fill color and font color
    header_cell = ws['A1']
    fill_color = header_cell.fill.start_color.rgb if header_cell.fill.start_color else None
    font_color = header_cell.font.color.rgb if header_cell.font.color else None
    print(f"Header fill color: {fill_color}")
    print(f"Header font color: {font_color}")

    # Get freeze panes value
    freeze_panes = ws.freeze_panes
    print(f"Freeze panes: {freeze_panes}")

    # Get max row
    print(f"Max row: {ws.max_row}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build ASR xlsx file")
    parser.add_argument("--year", type=int, required=True, help="Year for the xlsx file")

    args = parser.parse_args()

    build_xlsx(args.year)
    verify_xlsx(args.year)
