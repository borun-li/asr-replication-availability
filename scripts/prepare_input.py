#!/usr/bin/env python3
"""Prepare a volume's Block A worklist in ONE step.

    python3 scripts/prepare_input.py --vol 89      # -> input/asr_vol89.xlsx   (recommended)
    python3 scripts/prepare_input.py --year 2024   # -> input/asr_2024.xlsx    (by calendar year)

ASR is organized by **volume** (vol 89 = 2024, vol 90 = 2025, vol 91 = 2026 …). Because articles
publish online-first and are assigned to a volume later, `--vol N` fetches the two calendar years
around the volume and keeps only the articles whose Crossref `volume` == N, so a whole volume is
captured with no gaps or overlaps. Block A (title, authors, published-online date, url, volume,
issue, OnlineFirst) is filled; the coding columns are left empty.

Deterministic, no LLM. Needs an internet connection and `pip install requests openpyxl`.
Set CROSSREF_MAILTO to your email to use Crossref's faster polite pool (optional).
"""
import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_xlsx import build_xlsx                                  # noqa: E402
from format_xlsx import style                                     # noqa: E402
from crossref_fetch import fetch_all, online_date, authors, title_of  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
VOL_TO_YEAR = 1935  # ASR: year = volume + 1935  (vol 89 -> 2024)


def rows_from(items, vol_filter=None):
    seen, rows = set(), []
    for it in items:
        doi = (it.get("DOI") or "").strip().lower()
        if not doi or doi in seen:
            continue
        vol = str(it.get("volume") or "").strip()
        if vol_filter is not None and vol != str(vol_filter):
            continue
        seen.add(doi)
        rows.append({
            "title": title_of(it), "authors": authors(it), "online": online_date(it),
            "url": (it.get("resource", {}).get("primary", {}) or {}).get("URL", ""),
            "volume": vol, "issue": str(it.get("issue") or "").strip(),
        })
    rows.sort(key=lambda r: (r["online"], r["title"].lower()))
    return rows


def fill(out_path, rows):
    wb = load_workbook(out_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    def put(r, col, v):
        if col in idx:
            ws.cell(row=r, column=idx[col], value=v)

    for r, row in enumerate(rows, start=2):
        put(r, "title", row["title"])
        put(r, "author(s)", row["authors"])
        put(r, "published__online_date", row["online"])
        put(r, "article_url", row["url"])
        put(r, "volume", int(row["volume"]) if row["volume"].isdigit() else row["volume"] or None)
        put(r, "issue", int(row["issue"]) if row["issue"].isdigit() else row["issue"] or None)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Prepare a volume's Block A worklist in one step.")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--vol", type=int, help="ASR volume number, e.g. 89 (recommended)")
    g.add_argument("--year", type=int, help="calendar year, e.g. 2024")
    a = ap.parse_args()

    if a.vol is not None:
        years = [a.vol + VOL_TO_YEAR - 1, a.vol + VOL_TO_YEAR]   # prior + primary year
        out = ROOT / "input" / f"asr_vol{a.vol}.xlsx"
        sheet, vol_filter, label = f"asr_vol{a.vol}", a.vol, f"vol {a.vol}"
    else:
        years = [a.year]
        out = ROOT / "input" / f"asr_{a.year}.xlsx"
        sheet, vol_filter, label = f"asr_{a.year}", None, f"year {a.year}"

    print(f"Preparing Block A for {label} -> {out.name}")
    build_xlsx(out_path=out, sheet_name=sheet)

    items = []
    for y in years:
        print(f"  fetching Crossref year {y} …", flush=True)
        try:
            its, _ = fetch_all(y)
            items += its
        except Exception as e:
            print(f"  (year {y} skipped: {type(e).__name__}: {e})")

    rows = rows_from(items, vol_filter)
    if not rows:
        sys.exit(f"No articles found for {label}. (Check the volume/year, or your connection.)")
    fill(out, rows)
    style(out)
    print(f"\nDone -> {out}  ({len(rows)} articles; Block A complete, coding columns empty).")
    print("Next: code Block B — see README Scenario 2 (whole volume) or Scenario 3 (single article).")


if __name__ == "__main__":
    main()
