#!/usr/bin/env python3
"""Adds the OnlineFirst (Y/N) column after article_url and applies formatting
to /program/asr/input/asr_{YEAR}.xlsx.

OnlineFirst = Y when the SAGE article page showed no volume/issue (article is
published online ahead of assignment to an issue), else N.
"""

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
STRIPE = PatternFill("solid", fgColor="EEF2F8")        # very light blue
NO_FILL = PatternFill(fill_type=None)                  # explicit clear (odd rows)
OF_FILL = PatternFill("solid", fgColor="FFF2CC")       # highlight OnlineFirst=Y
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "title": 62, "author(s)": 34, "published__online_date": 15,
    "article_url": 46, "OnlineFirst (Y/N)": 12, "volume": 8, "issue": 7,
    "in_scope(Y/NA)": 12, "qualitative(Y/N)": 13,
    "data(Y/N)": 10, "code(Y/N)": 10, "data + code": 12, "neither": 10,
    "data_gated(Y/N)": 13, "data_source / apply_at": 26,
    "package_location": 26, "path_to_package": 34, "coverage_checked": 14,
    "notes": 40,
}
CENTERED = {"published__online_date", "OnlineFirst (Y/N)", "volume", "issue",
            "in_scope(Y/NA)", "qualitative(Y/N)", "data(Y/N)",
            "code(Y/N)", "data + code", "neither", "data_gated(Y/N)",
            "coverage_checked"}


def style(path, sheet=None):
    """Add the OnlineFirst column + apply the standard formatting to a workbook, in place."""
    wb = load_workbook(path)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    headers = [c.value for c in ws[1]]

    # --- 1. insert OnlineFirst (Y/N) right after article_url -----------------
    if "OnlineFirst (Y/N)" not in headers:
        at = headers.index("article_url") + 2
        ws.insert_cols(at)
        ws.cell(row=1, column=at, value="OnlineFirst (Y/N)")
        headers = [c.value for c in ws[1]]

    idx = {h: i + 1 for i, h in enumerate(headers)}
    of_col, vol_col, iss_col = (idx["OnlineFirst (Y/N)"], idx["volume"],
                                idx["issue"])

    online_first = []
    for r in range(2, ws.max_row + 1):
        no_issue = (ws.cell(row=r, column=vol_col).value in (None, "")
                    and ws.cell(row=r, column=iss_col).value in (None, ""))
        ws.cell(row=r, column=of_col, value="Y" if no_issue else "N")
        if no_issue:
            online_first.append((r, ws.cell(row=r, column=1).value))

    # --- 2. formatting -------------------------------------------------------
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(
            headers[c - 1], 16)
    ws.row_dimensions[1].height = 34

    for r in range(2, ws.max_row + 1):
        is_of = ws.cell(row=r, column=of_col).value == "Y"
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            centered = headers[c - 1] in CENTERED
            cell.alignment = Alignment(
                horizontal="center" if centered else "left",
                vertical="top", wrap_text=headers[c - 1] in
                ("title", "author(s)", "notes", "path_to_package"))
            # set the fill on EVERY body cell so re-styling never leaves a stale
            # highlight behind (odd rows are explicitly cleared, not skipped)
            cell.fill = STRIPE if r % 2 == 0 else NO_FILL
        if is_of:
            ws.cell(row=r, column=of_col).fill = OF_FILL
            ws.cell(row=r, column=of_col).font = Font(bold=True, color="9C5700")
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{ws.max_row}"
    ws.sheet_view.zoomScale = 110

    wb.save(path)
    return online_first


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, required=True)
    args = ap.parse_args()
    path = ROOT / "input" / f"asr_{args.year}.xlsx"
    of = style(path)
    print(f"Saved {path} — OnlineFirst = Y on {len(of)} rows.")


if __name__ == "__main__":
    main()
