#!/usr/bin/env python3
"""
Task 2 - crossref_fetch.py  (NO LLM; deterministic)

Fetches all American Sociological Review (ISSN 0003-1224) journal-articles whose
online publication date falls within {YEAR}, saves the raw Crossref JSON, and
fills ONLY columns 1-4 of /program/asr/input/asr_{YEAR}.xlsx:

    title                  <- title[0]
    author(s)              <- join(author[], "{given} {family}", "; ")
    published__online_date <- published-online.date-parts[0] as YYYY-MM-DD
    article_url            <- resource.primary.URL

HARD RULES: deduplicate on DOI; leave all other columns blank.
"""

import argparse
import json
import sys
import time
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]          # /Users/apple/program/asr
CACHE_DIR = ROOT / "cache"
INPUT_DIR = ROOT / "input"

ISSN = "0003-1224"
# Crossref "polite pool" contact. Set CROSSREF_MAILTO to your own email for faster/more reliable
# service; left unset it queries the anonymous pool. Never hard-code a personal address here.
import os
MAILTO = os.environ.get("CROSSREF_MAILTO", "").strip()
API = "https://api.crossref.org/works"
ROWS = 1000


def fetch_all(year: int):
    """Paginate Crossref via message.next-cursor until items is empty."""
    filters = ",".join([
        f"issn:{ISSN}",
        "type:journal-article",
        f"from-online-pub-date:{year}-01-01",
        f"until-online-pub-date:{year}-12-31",
    ])
    session = requests.Session()
    ua = "asr-replication-audit/1.0" + (f" (mailto:{MAILTO})" if MAILTO else "")
    session.headers.update({"User-Agent": ua})

    cursor = "*"
    items = []
    pages = 0
    total = None
    while True:
        params = {
            "filter": filters,
            "rows": ROWS,
            "cursor": cursor,
        }
        if MAILTO:
            params["mailto"] = MAILTO
        url = f"{API}?{urlencode(params)}"
        resp = session.get(url, timeout=60)
        resp.raise_for_status()
        msg = resp.json()["message"]
        batch = msg.get("items", [])
        pages += 1
        if total is None:
            total = msg.get("total-results")
            print(f"Crossref total-results: {total}")
        print(f"  page {pages}: {len(batch)} items")
        if not batch:
            break
        items.extend(batch)
        cursor = msg.get("next-cursor")
        if not cursor:
            break
        time.sleep(1)
    return items, total


def online_date(item):
    parts = (item.get("published-online") or {}).get("date-parts") or [[]]
    p = parts[0] if parts else []
    if not p:
        return ""
    y = p[0]
    m = p[1] if len(p) > 1 else 1
    d = p[2] if len(p) > 2 else 1
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"


def authors(item):
    out = []
    for a in item.get("author") or []:
        given = (a.get("given") or "").strip()
        family = (a.get("family") or "").strip()
        name = " ".join(x for x in (given, family) if x).strip()
        if not name:
            name = (a.get("name") or "").strip()
        if name:
            out.append(name)
    return "; ".join(out)


def title_of(item):
    t = item.get("title") or []
    return (t[0] if t else "").strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, required=True)
    args = ap.parse_args()
    year = args.year

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    raw_path = CACHE_DIR / f"crossref_{year}.json"

    items, total = fetch_all(year)

    # Save raw JSON FIRST.
    with raw_path.open("w", encoding="utf-8") as fh:
        json.dump({"query": {"issn": ISSN, "year": year, "total_results": total},
                   "items": items}, fh, ensure_ascii=False, indent=2)
    print(f"Raw JSON saved -> {raw_path} ({len(items)} items)")

    # Deduplicate on DOI (case-insensitive), keep first occurrence.
    seen = set()
    rows = []
    for it in items:
        doi = (it.get("DOI") or "").strip().lower()
        if not doi or doi in seen:
            continue
        seen.add(doi)
        rows.append({
            "title": title_of(it),
            "authors": authors(it),
            "online": online_date(it),
            "url": (it.get("resource", {}).get("primary", {}) or {}).get("URL", ""),
            "doi": doi,
        })
    dropped = len(items) - len(rows)
    print(f"After DOI dedup: {len(rows)} rows ({dropped} duplicates/no-DOI dropped)")

    # Sort by online date, then title, for stable ordering.
    rows.sort(key=lambda r: (r["online"], r["title"].lower()))

    xlsx = INPUT_DIR / f"asr_{year}.xlsx"
    if not xlsx.exists():
        sys.exit(f"ERROR: {xlsx} not found - run Task 1 (build_xlsx.py) first.")
    wb = load_workbook(xlsx)
    ws = wb[f"asr_{year}"] if f"asr_{year}" in wb.sheetnames else wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for col in ("title", "author(s)", "published__online_date", "article_url"):
        if col not in idx:
            sys.exit(f"ERROR: column '{col}' missing from {xlsx}")

    # Clear any pre-existing data rows (headers only workbook expected).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=idx["title"], value=row["title"])
        ws.cell(row=r, column=idx["author(s)"], value=row["authors"])
        ws.cell(row=r, column=idx["published__online_date"], value=row["online"])
        ws.cell(row=r, column=idx["article_url"], value=row["url"])
        # All other columns intentionally left blank.

    wb.save(xlsx)
    print(f"Wrote {len(rows)} rows -> {xlsx}")


if __name__ == "__main__":
    main()
