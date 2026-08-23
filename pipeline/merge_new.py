#!/usr/bin/env python3
"""Merge coded new articles into the dataset, by volume (Scenario 3, Step 3).

Reads a coded worklist (default `input/new_articles.csv`, the 19-column ASR layout) and folds each
**coded** row into the matching `output/asr_vol<N>_result.csv` (and the `.xlsx` if openpyxl is
installed). Safe merge:
  - only rows that are actually coded (in_scope filled) are added — uncoded rows are skipped,
  - rows whose DOI already exists in that volume's dataset are skipped as duplicates,
  - `data + code` / `neither` are recomputed from `data` / `code`.

Usage:
    python3 pipeline/merge_new.py                     # merge input/new_articles.csv
    python3 pipeline/merge_new.py path/to/coded.csv   # a different source
    python3 pipeline/merge_new.py --dry-run           # show what would happen, change nothing
"""
import csv
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, os.pardir)
OUT = os.path.join(ROOT, "output")
DOI_RE = re.compile(r"(10\.1177/[0-9A-Za-z._]+)")


def doi_of(url):
    m = DOI_RE.search(url or "")
    return (m.group(1).rstrip("/").lower()) if m else ""


def col(fieldnames, sub):
    for h in fieldnames or []:
        if h and sub in str(h).lower():
            return h
    return None


def main():
    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        return 0
    dry = "--dry-run" in args
    srcs = [a for a in args if a != "--dry-run"]
    src = srcs[0] if srcs else os.path.join(ROOT, "input", "new_articles.csv")
    if not os.path.exists(src):
        print(f"ERROR: source worklist not found: {src}", file=sys.stderr)
        return 2

    with open(src, newline="", encoding="utf-8") as f:
        rd = csv.DictReader(f)
        src_cols = rd.fieldnames
        new_rows = list(rd)
    c_url = col(src_cols, "article_url")
    c_vol = col(src_cols, "volume")
    c_ins = col(src_cols, "in_scope")
    c_data = col(src_cols, "data(y")
    c_code = col(src_cols, "code(y")
    c_dc = col(src_cols, "data + code")
    c_nei = col(src_cols, "neither")

    # bucket coded rows by target volume file
    buckets, skipped = {}, []
    for r in new_rows:
        doi = doi_of(r.get(c_url))
        vol = str(r.get(c_vol) or "").strip()
        if not (r.get(c_ins) or "").strip():
            skipped.append((doi or "?", "not coded yet (in_scope empty)"))
            continue
        if not vol:
            skipped.append((doi or "?", "no volume — cannot route"))
            continue
        target = os.path.join(OUT, f"asr_vol{vol}_result.csv")
        buckets.setdefault(target, []).append((doi, r))

    total_add = 0
    for target, rows in sorted(buckets.items()):
        if not os.path.exists(target):
            print(f"  NOTE: {os.path.basename(target)} does not exist yet — it will be created.")
            existing_dois, tgt_cols, tgt_rows = set(), src_cols, []
        else:
            with open(target, newline="", encoding="utf-8") as f:
                rd = csv.DictReader(f)
                tgt_cols = rd.fieldnames
                tgt_rows = list(rd)
            tcu = col(tgt_cols, "article_url")
            existing_dois = {doi_of(r.get(tcu)) for r in tgt_rows}
        to_add = []
        for doi, r in rows:
            if doi and doi in existing_dois:
                skipped.append((doi, f"duplicate in {os.path.basename(target)}"))
                continue
            row = {c: (r.get(c) or "") for c in tgt_cols}
            d = str(r.get(c_data) or "").strip().upper()
            co = str(r.get(c_code) or "").strip().upper()
            if c_dc and col(tgt_cols, "data + code"):
                row[col(tgt_cols, "data + code")] = "Y" if (d == "Y" and co == "Y") else "N"
            if c_nei and col(tgt_cols, "neither"):
                row[col(tgt_cols, "neither")] = "Y" if (d == "N" and co == "N") else "N"
            to_add.append(row)
        print(f"  {os.path.basename(target)}: +{len(to_add)} row(s)")
        total_add += len(to_add)
        if not dry and to_add:
            write_header = not os.path.exists(target)
            with open(target, "a", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=tgt_cols)
                if write_header:
                    w.writeheader()
                for row in to_add:
                    w.writerow(row)
            _sync_xlsx(target, tgt_cols, to_add)

    print(f"\n{'[dry-run] would add' if dry else 'Added'} {total_add} row(s).")
    if skipped:
        print(f"Skipped {len(skipped)}:")
        for k, why in skipped[:20]:
            print(f"  - {k}: {why}")
    if not dry and total_add:
        print("Verify with:  python3 pipeline/reproduce_table.py")
    return 0


def _sync_xlsx(csv_target, cols, rows):
    xlsx = csv_target[:-4] + ".xlsx"
    if not os.path.exists(xlsx):
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("    (xlsx not updated — `pip install openpyxl` to sync it; the CSV is authoritative.)")
        return
    wb = load_workbook(xlsx)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    for r in rows:
        ws.append([r.get(h, "") for h in hdr])
    wb.save(xlsx)


if __name__ == "__main__":
    raise SystemExit(main())
